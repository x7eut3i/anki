import asyncio
import logging

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, UploadFile, File
from sqlmodel import Session, select

from app.auth import get_current_user
from app.database import get_session, engine
from app.models.ai_config import AIConfig
from app.models.user import User
from app.schemas.ai import (
    AIConfigUpdate,
    AIConfigResponse,
    AIConfigCreate,
    AITestRequest,
    AITestResponse,
    AIExplainRequest,
    AIExplainResponse,
    AIChatRequest,
    AIChatResponse,
    AIGenerateCardsRequest,
    AIGenerateCardsResponse,
    AIUsageResponse,
    AIBatchEnrichRequest,
    AIBatchEnrichResponse,
)
from app.services.ai_service import AIService
from app.services.prompt_loader import get_prompt, get_prompt_model
from app.routers.ai_jobs import create_job, update_job_status

logger = logging.getLogger("anki.ai")

router = APIRouter(prefix="/api/ai", tags=["ai"])


def _config_to_response(config: AIConfig) -> AIConfigResponse:
    return AIConfigResponse(
        id=config.id,
        name=getattr(config, "name", "默认") or "默认",
        api_base_url=config.api_base_url,
        api_key_set=bool(config.api_key),
        model=config.model,
        model_pipeline=config.model_pipeline or "",
        model_reading=config.model_reading or "",
        max_daily_calls=config.max_daily_calls,
        import_batch_size=getattr(config, "import_batch_size", 30) or 30,
        import_concurrency=getattr(config, "import_concurrency", 3) or 3,
        max_tokens=getattr(config, "max_tokens", 8192) or 8192,
        temperature=getattr(config, "temperature", 0.3) or 0.3,
        max_retries=getattr(config, "max_retries", 3) or 3,
        is_enabled=config.is_enabled,
        is_active=getattr(config, "is_active", True),
        auto_explain_wrong=config.auto_explain_wrong,
        auto_generate_mnemonics=config.auto_generate_mnemonics,
        auto_generate_related=config.auto_generate_related,
        updated_at=config.updated_at,
    )


def _get_active_config(session: Session, user_id: int) -> AIConfig | None:
    """Get the active AI config for a user."""
    # Try to find active config first
    config = session.exec(
        select(AIConfig).where(
            AIConfig.user_id == user_id,
            AIConfig.is_active == True
        )
    ).first()
    if not config:
        # Fall back to any config
        config = session.exec(
            select(AIConfig).where(AIConfig.user_id == user_id)
        ).first()
    return config


@router.get("/configs", response_model=list[AIConfigResponse])
def list_ai_configs(
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """List all AI configs for the current user."""
    configs = session.exec(
        select(AIConfig).where(AIConfig.user_id == current_user.id)
    ).all()
    if not configs:
        # Create a default config
        config = AIConfig(user_id=current_user.id, name="默认", is_active=True)
        from app.main import _load_ai_config_file
        defaults = _load_ai_config_file()
        if defaults and defaults.get("api_key"):
            config.api_base_url = defaults["api_base_url"]
            config.api_key = defaults["api_key"]
            config.model = defaults["model"]
            config.max_daily_calls = defaults["max_daily_calls"]
            config.is_enabled = True
        session.add(config)
        session.commit()
        session.refresh(config)
        configs = [config]
    return [_config_to_response(c) for c in configs]


@router.post("/configs", response_model=AIConfigResponse, status_code=201)
def create_ai_config(
    data: AIConfigCreate,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """Create a new AI config profile."""
    config = AIConfig(
        user_id=current_user.id,
        name=data.name,
        is_active=False,
        api_base_url=data.api_base_url or "https://api.openai.com/v1",
        api_key=data.api_key,
        model=data.model or "gpt-4o-mini",
        model_pipeline=data.model_pipeline,
        model_reading=data.model_reading,
        max_daily_calls=data.max_daily_calls,
        import_batch_size=data.import_batch_size,
        import_concurrency=data.import_concurrency,
    )
    session.add(config)
    session.commit()
    session.refresh(config)
    return _config_to_response(config)


@router.post("/configs/{config_id}/activate", response_model=AIConfigResponse)
def activate_ai_config(
    config_id: int,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """Set a config as the active one (deactivates others)."""
    config = session.get(AIConfig, config_id)
    if not config or config.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="配置不存在")

    # Deactivate all others
    all_configs = session.exec(
        select(AIConfig).where(AIConfig.user_id == current_user.id)
    ).all()
    for c in all_configs:
        c.is_active = c.id == config_id
        session.add(c)

    session.commit()
    session.refresh(config)
    return _config_to_response(config)


@router.put("/configs/{config_id}/rename", response_model=AIConfigResponse)
def rename_ai_config(
    config_id: int,
    data: dict,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """Rename an AI config profile."""
    config = session.get(AIConfig, config_id)
    if not config or config.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="配置不存在")

    new_name = data.get("name", "").strip()
    if not new_name:
        raise HTTPException(status_code=400, detail="名称不能为空")

    config.name = new_name
    session.add(config)
    session.commit()
    session.refresh(config)
    return _config_to_response(config)


@router.delete("/configs/{config_id}")
def delete_ai_config(
    config_id: int,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """Delete an AI config. Cannot delete the last one."""
    configs = session.exec(
        select(AIConfig).where(AIConfig.user_id == current_user.id)
    ).all()
    if len(configs) <= 1:
        raise HTTPException(status_code=400, detail="无法删除最后一个配置")
    config = session.get(AIConfig, config_id)
    if not config or config.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="配置不存在")

    was_active = getattr(config, "is_active", False)
    session.delete(config)
    session.commit()

    # If deleted the active one, activate the first remaining
    if was_active:
        remaining = session.exec(
            select(AIConfig).where(AIConfig.user_id == current_user.id)
        ).first()
        if remaining:
            remaining.is_active = True
            session.add(remaining)
            session.commit()

    return {"ok": True}


@router.get("/config", response_model=AIConfigResponse)
def get_ai_config(
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    config = _get_active_config(session, current_user.id)

    if not config:
        config = AIConfig(user_id=current_user.id, name="默认", is_active=True)
        # Auto-populate from ai_config.json if available
        from app.main import _load_ai_config_file
        defaults = _load_ai_config_file()
        if defaults and defaults.get("api_key"):
            config.api_base_url = defaults["api_base_url"]
            config.api_key = defaults["api_key"]
            config.model = defaults["model"]
            config.max_daily_calls = defaults["max_daily_calls"]
            config.is_enabled = True
        session.add(config)
        session.commit()
        session.refresh(config)

    return _config_to_response(config)


@router.put("/config", response_model=AIConfigResponse)
def update_ai_config(
    data: AIConfigUpdate,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    config = _get_active_config(session, current_user.id)

    if not config:
        config = AIConfig(user_id=current_user.id, name="默认", is_active=True)

    update_data = data.model_dump(exclude_unset=True)
    # Never overwrite api_key with empty string
    if "api_key" in update_data and not update_data["api_key"]:
        del update_data["api_key"]
    for key, value in update_data.items():
        setattr(config, key, value)

    session.add(config)
    session.commit()
    session.refresh(config)

    return _config_to_response(config)


@router.post("/test-connection", response_model=AITestResponse)
async def test_ai_connection(
    data: AITestRequest,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    # If api_key is empty, use the stored key from user's config
    api_key = data.api_key
    api_base_url = data.api_base_url
    model_name = data.model
    if not api_key:
        config = _get_active_config(session, current_user.id)
        if config and config.api_key:
            api_key = config.api_key
            if not api_base_url:
                api_base_url = config.api_base_url
            if not model_name:
                model_name = config.model
        else:
            return AITestResponse(success=False, message="未找到已保存的API密钥，请输入密钥")

    logger.debug("Testing AI connection to %s with model %s", api_base_url, model_name)
    result = await AIService.test_connection(
        api_base_url=api_base_url,
        api_key=api_key,
        model=model_name,
    )
    logger.debug("Connection test result: success=%s, message=%s", result.get('success'), result.get('message'))
    return AITestResponse(**result)


@router.post("/models")
async def list_ai_models(
    data: AITestRequest,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """Fetch available models from the AI provider."""
    # If api_key is empty, use the stored key
    api_key = data.api_key
    api_base_url = data.api_base_url
    if not api_key:
        config = _get_active_config(session, current_user.id)
        if config and config.api_key:
            api_key = config.api_key
            if not api_base_url:
                api_base_url = config.api_base_url

    logger.debug("Fetching models from %s", api_base_url)
    result = await AIService.list_models(
        api_base_url=api_base_url,
        api_key=api_key,
    )
    logger.debug("Models result: success=%s, count=%d", result.get('success'), len(result.get('models', [])))
    return result


@router.post("/explain", response_model=AIExplainResponse)
async def explain_card(
    data: AIExplainRequest,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    from app.models.card import Card
    card = session.get(Card, data.card_id)
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")

    ai = AIService(session, current_user.id)
    if not ai.is_available():
        raise HTTPException(status_code=503, detail=ai.get_unavailable_reason())

    result = await ai.explain_card(card.front, card.back, data.user_answer)
    return AIExplainResponse(**result)


@router.post("/chat", response_model=AIChatResponse)
async def chat_with_tutor(
    data: AIChatRequest,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    ai = AIService(session, current_user.id)
    if not ai.is_available():
        raise HTTPException(status_code=503, detail=ai.get_unavailable_reason())

    card_context = ""
    if data.card_id:
        from app.models.card import Card
        card = session.get(Card, data.card_id)
        if card:
            card_context = f"题目: {card.front}\n答案: {card.back}"

    result = await ai.chat_tutor(data.message, data.history, card_context)
    return AIChatResponse(**result)


@router.post("/generate", response_model=AIGenerateCardsResponse)
async def generate_cards(
    data: AIGenerateCardsRequest,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    ai = AIService(session, current_user.id)
    if not ai.is_available():
        raise HTTPException(status_code=503, detail=ai.get_unavailable_reason())

    # Get category name and all available categories
    from app.models.category import Category
    cat_name = ""
    if data.category_id:
        cat = session.get(Category, data.category_id)
        if cat:
            cat_name = cat.name

    all_cats = session.exec(select(Category).where(Category.is_active == True)).all()
    available_categories = [c.name for c in all_cats]

    cards = await ai.generate_cards_from_text(
        text=data.text,
        category_name=cat_name,
        card_type=data.card_type,
        count=data.count,
        available_categories=available_categories,
    )
    return AIGenerateCardsResponse(cards=cards, tokens_used=0)


@router.get("/usage", response_model=AIUsageResponse)
def get_usage(
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    ai = AIService(session, current_user.id)
    stats = ai.get_usage_stats()
    return AIUsageResponse(**stats)


@router.post("/batch-enrich")
async def batch_enrich_cards(
    data: AIBatchEnrichRequest,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """Batch-enrich cards with AI-generated distractors, meta_info, etc."""
    from app.models.card import Card
    import json as _json

    ai = AIService(session, current_user.id)
    if not ai.is_available():
        raise HTTPException(status_code=503, detail=ai.get_unavailable_reason())

    # Collect cards to enrich
    if data.card_ids:
        cards_to_enrich = []
        for cid in data.card_ids:
            card = session.get(Card, cid)
            if card:
                cards_to_enrich.append(card)
    elif data.deck_id:
        from sqlmodel import select as sel
        cards_to_enrich = list(session.exec(
            sel(Card).where(Card.deck_id == data.deck_id)
        ).all())
    else:
        raise HTTPException(status_code=400, detail="Provide card_ids or deck_id")

    if not cards_to_enrich:
        return AIBatchEnrichResponse(message="No cards found")

    total = len(cards_to_enrich)
    enriched = 0
    skipped = 0
    errors = 0

    # Process in batches
    batch_size = data.batch_size
    for i in range(0, total, batch_size):
        batch = cards_to_enrich[i:i + batch_size]

        # Build batch prompt
        cards_text = []
        for idx, card in enumerate(batch):
            has_distractors = bool(card.distractors and card.distractors.strip() and card.distractors != "[]")
            has_meta = bool(card.meta_info and card.meta_info.strip() and card.meta_info not in ("{}", ""))
            if has_distractors and has_meta:
                skipped += 1
                continue
            cards_text.append({
                "index": idx,
                "card_id": card.id,
                "front": card.front[:200],
                "back": card.back[:200],
                "explanation": (card.explanation or "")[:100],
                "needs_distractors": not has_distractors,
                "needs_meta": not has_meta,
            })

        if not cards_text:
            continue

        # Build AI prompt
        prompt_cards = _json.dumps(cards_text, ensure_ascii=False)
        default_enrich_prompt = (
            "你是出题专家。请为以下卡片补充缺失的信息。\n"
            "对于needs_distractors=true的卡片，生成3个合理的干扰项(distractors)。\n"
            "对于needs_meta=true的卡片，生成meta_info，包含：\n"
            '  - knowledge: {key_points, synonyms, antonyms, related, memory_tips, golden_quotes, formal_terms, essay_material}\n'
            '  - exam_focus: {difficulty: "easy/medium/hard", frequency: "high/medium/low"}\n'
            '  - facts: {关键事实}\n'
            '注意：meta_info中不要包含distractors字段，干扰项只放在顶层distractors字段中。\n\n'
            "回复JSON数组格式，每项包含card_id和补充的字段:\n"
            '[{"card_id": 1, "distractors": ["A","B","C"], "meta_info": {...}}, ...]'
        )
        messages = [
            {
                "role": "system",
                "content": get_prompt(session, "batch_enrich", default_enrich_prompt),
            },
            {"role": "user", "content": prompt_cards},
        ]

        try:
            result = await ai.chat_completion(messages, temperature=0.3, feature="batch_enrich")
            content = result["content"]
            if "```json" in content:
                content = content.split("```json")[1].split("```")[0]
            elif "```" in content:
                content = content.split("```")[1].split("```")[0]
            enrichments = _json.loads(content)

            # Apply enrichments
            card_map = {c.id: c for c in batch}
            for item in enrichments:
                cid = item.get("card_id")
                card = card_map.get(cid)
                if not card:
                    continue
                if "distractors" in item and item["distractors"]:
                    card.distractors = _json.dumps(item["distractors"], ensure_ascii=False)
                if "meta_info" in item and item["meta_info"]:
                    meta = item["meta_info"]
                    if isinstance(meta, dict):
                        meta = _json.dumps(meta, ensure_ascii=False)
                    card.meta_info = meta
                session.add(card)
                enriched += 1

            session.commit()
        except Exception as e:
            logger.error("Batch enrich error: %s", e)
            errors += len(cards_text)

    return AIBatchEnrichResponse(
        total=total,
        enriched=enriched,
        skipped=skipped,
        errors=errors,
        message=f"已处理 {total} 张卡片：{enriched} 张已补充，{skipped} 张已跳过",
    )


@router.post("/complete-cards")
async def complete_cards(
    data: dict,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """AI-complete cards: generate back/explanation/distractors/meta from front text.

    Uses the same CARD_SYSTEM_PROMPT as import and article card generation
    to ensure consistent card quality and format across all scenarios.

    Expects: { "cards": [{"front": "...", "category": "..."}], "deck_id": int }
    Returns: { "cards": [{"front": "...", "back": "...", "explanation": "...", "distractors": [...], "meta_info": "...", "tags": "..."}] }
    """
    import json as _json
    import time as _time
    from app.services.prompts import CARD_SYSTEM_PROMPT
    from app.models.category import Category

    ai = AIService(session, current_user.id)
    if not ai.is_available():
        raise HTTPException(status_code=503, detail=ai.get_unavailable_reason())

    cards_input = data.get("cards", [])
    if not cards_input:
        raise HTTPException(status_code=400, detail="No cards provided")

    # Get available categories
    all_cats = session.exec(select(Category).where(Category.is_active == True)).all()
    cat_list = "、".join(c.name for c in all_cats)

    # Build prompt using the same format as pipeline/import
    cards_text = []
    for i, c in enumerate(cards_input):
        cards_text.append({
            "index": i,
            "front": c.get("front", ""),
            "category": c.get("category", ""),
        })

    user_prompt = (
        f"请为以下卡片生成完整的内容（JSON数组）。\n"
        f"每张卡片已有front（题目）和category（分类），请根据system prompt中的格式要求补充所有字段：\n"
        f"back、explanation、distractors（3个）、tags、meta_info（含knowledge、exam_focus、alternate_questions、facts）。\n\n"
        f"可用分类：{cat_list}\n\n"
        f"待补全的卡片：\n{_json.dumps(cards_text, ensure_ascii=False)}\n\n"
        f"回复纯JSON数组，每项包含index字段和所有补全的字段。"
    )

    messages = [
        {"role": "system", "content": get_prompt(session, "card_system", CARD_SYSTEM_PROMPT)},
        {"role": "user", "content": user_prompt},
    ]

    try:
        t0 = _time.time()
        result = await ai.chat_completion(messages, feature="complete_cards")
        elapsed_ms = int((_time.time() - t0) * 1000)

        content = result["content"]
        tokens_used = result.get("tokens_used", 0)

        # Parse JSON
        if "```json" in content:
            content = content.split("```json")[1].split("```")[0]
        elif "```" in content:
            content = content.split("```")[1].split("```")[0]

        import re
        content = re.sub(r",\s*([}\]])", r"\1", content)  # Remove trailing commas
        parsed = _json.loads(content)

        # Handle pipeline-style response with {cards: [...]}
        if isinstance(parsed, dict):
            completions = parsed.get("cards", [])
        elif isinstance(parsed, list):
            completions = parsed
        else:
            completions = []
    except Exception as e:
        logger.error("Complete cards AI error: %s", e)
        # Graceful degradation: return original cards without AI enrichment
        return {"cards": [{"front": c.get("front", ""), "back": c.get("back", ""), "explanation": "", "distractors": "", "meta_info": "", "tags": "", "category": c.get("category", "")} for c in cards_input], "completed": 0, "error": f"AI 补全失败: {str(e)[:200]}"}

    if not isinstance(completions, list):
        return {"cards": [{"front": c.get("front", ""), "back": c.get("back", ""), "explanation": "", "distractors": "", "meta_info": "", "tags": "", "category": c.get("category", "")} for c in cards_input], "completed": 0, "error": "AI 返回格式错误"}

    # Merge completions back into input cards
    result_cards = []
    comp_map = {item.get("index", i): item for i, item in enumerate(completions)}
    for i, c in enumerate(cards_input):
        comp = comp_map.get(i, {})
        distractors = comp.get("distractors", [])
        if isinstance(distractors, list):
            distractors_str = _json.dumps(distractors, ensure_ascii=False)
        else:
            distractors_str = str(distractors) if distractors else ""

        meta_info = comp.get("meta_info", "")
        if isinstance(meta_info, dict):
            meta_info = _json.dumps(meta_info, ensure_ascii=False)

        result_cards.append({
            "front": comp.get("front", c.get("front", "")),
            "back": comp.get("back", c.get("back", "")),
            "explanation": comp.get("explanation", c.get("explanation", "")),
            "distractors": distractors_str,
            "meta_info": meta_info,
            "tags": comp.get("tags", c.get("tags", "")),
            "category": comp.get("category", c.get("category", "")),
        })

    return {"cards": result_cards, "completed": len(completions)}


@router.post("/smart-import")
async def smart_import(
    deck_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """AI-powered smart import: auto-parse any file format into flashcards."""
    from app.models.card import Card
    from app.models.deck import Deck
    import json as _json

    ai = AIService(session, current_user.id)
    if not ai.is_available():
        raise HTTPException(status_code=503, detail=ai.get_unavailable_reason())

    deck = session.get(Deck, deck_id)
    if not deck:
        raise HTTPException(status_code=404, detail="Deck not found")

    # Read file content
    file_bytes = await file.read()
    filename = file.filename or "unknown"

    # Try to detect content type and parse
    content_text = ""
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        import openpyxl
        import io
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        content_text = "\n".join(["\t".join(str(c) if c else "" for c in row) for row in rows[:100]])
        wb.close()
    else:
        try:
            content_text = file_bytes.decode("utf-8")
        except UnicodeDecodeError:
            content_text = file_bytes.decode("gbk", errors="replace")

    # Limit content to prevent token overflow
    if len(content_text) > 8000:
        content_text = content_text[:8000] + "\n...(truncated)"

    # Ask AI to parse and generate cards — use same prompt as pipeline
    from app.services.prompts import CARD_SYSTEM_PROMPT, make_import_user_prompt
    from app.models.category import Category

    all_cats = session.exec(select(Category).where(Category.is_active == True)).all()
    cat_list = "、".join(c.name for c in all_cats)
    cat_map = {c.name: c for c in all_cats}

    user_prompt = make_import_user_prompt(
        filename=filename,
        batch_text=content_text,
        category_list=cat_list,
    )

    messages = [
        {
            "role": "system",
            "content": get_prompt(session, "card_system", CARD_SYSTEM_PROMPT),
        },
        {"role": "user", "content": user_prompt},
    ]

    try:
        result = await ai.chat_completion(messages, temperature=0.2, feature="smart_import")
        content = result["content"]
        if "```json" in content:
            content = content.split("```json")[1].split("```")[0]
        elif "```" in content:
            content = content.split("```")[1].split("```")[0]
        import re
        content = re.sub(r",\s*([}\]])", r"\1", content)
        cards_data = _json.loads(content)
    except Exception as e:
        logger.error("Smart import AI parsing error: %s", e)
        raise HTTPException(status_code=422, detail=f"AI 解析文件失败: {str(e)[:200]}")

    if not isinstance(cards_data, list):
        # Pipeline prompt may return {"cards": [...]}
        if isinstance(cards_data, dict) and "cards" in cards_data:
            cards_data = cards_data["cards"]
        else:
            raise HTTPException(status_code=422, detail="AI 返回的不是有效的卡片数组")

    created = 0
    errors_list = []
    for i, item in enumerate(cards_data):
        try:
            distractors = item.get("distractors", [])
            if isinstance(distractors, list):
                distractors_str = _json.dumps(distractors, ensure_ascii=False)
            else:
                distractors_str = str(distractors) if distractors else ""

            meta_info = item.get("meta_info", "")
            if isinstance(meta_info, dict):
                meta_info = _json.dumps(meta_info, ensure_ascii=False)

            # Try to match category
            cat_name = item.get("category", "")
            category = cat_map.get(cat_name)
            category_id = category.id if category else None

            card = Card(
                deck_id=deck_id,
                category_id=category_id,
                front=item.get("front", ""),
                back=item.get("back", ""),
                explanation=item.get("explanation", ""),
                distractors=distractors_str,
                tags=item.get("tags", ""),
                meta_info=meta_info,
                is_ai_generated=True,
            )
            session.add(card)
            created += 1
        except Exception as e:
            errors_list.append(f"Card {i}: {str(e)}")

    deck.card_count += created
    session.add(deck)
    session.commit()

    return {
        "created": created,
        "total_parsed": len(cards_data),
        "errors": errors_list,
        "message": f"AI 智能导入完成：解析 {len(cards_data)} 张，成功导入 {created} 张",
    }


@router.post("/ingest")
async def trigger_ingestion(
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """Trigger article ingestion pipeline (admin only)."""
    if not current_user.is_admin:
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Admin only")

    logger.info("Ingestion pipeline triggered by user %s", current_user.username)
    from app.services.ingestion_service import IngestionService
    service = IngestionService(session, current_user.id)
    result = await service.run_daily_ingestion()
    logger.debug("Ingestion result: %s", result)
    return result


# ── Async Job Wrappers ──────────────────────────────────────────────────

def _run_async(coro):
    """Run an async coroutine in a new event loop (for background threads)."""
    loop = asyncio.new_event_loop()
    try:
        return loop.run_until_complete(coro)
    finally:
        loop.close()


def _bg_smart_import(job_id: int, user_id: int, deck_id: int, file_bytes: bytes, filename: str):
    """Background task for smart import."""
    from sqlmodel import Session as SyncSession
    from app.models.card import Card
    from app.models.deck import Deck
    import json as _json

    update_job_status(job_id, "running", progress=10)

    try:
        with SyncSession(engine) as session:
            ai = AIService(session, user_id)
            if not ai.is_available():
                update_job_status(job_id, "failed", error_message=ai.get_unavailable_reason())
                return

            deck = session.get(Deck, deck_id)
            if not deck:
                update_job_status(job_id, "failed", error_message="牌组不存在")
                return

            # Parse file content
            content_text = ""
            if filename.endswith((".xlsx", ".xls")):
                import openpyxl, io
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                content_text = "\n".join(["\t".join(str(c) if c else "" for c in row) for row in rows[:100]])
                wb.close()
            else:
                try:
                    content_text = file_bytes.decode("utf-8")
                except UnicodeDecodeError:
                    content_text = file_bytes.decode("gbk", errors="replace")

            if len(content_text) > 8000:
                content_text = content_text[:8000] + "\n...(truncated)"

            update_job_status(job_id, "running", progress=30)

            from app.services.prompts import CARD_SYSTEM_PROMPT, make_import_user_prompt
            from app.models.category import Category

            all_cats = session.exec(select(Category).where(Category.is_active == True)).all()
            cat_list = "、".join(c.name for c in all_cats)
            cat_map = {c.name: c for c in all_cats}

            user_prompt = make_import_user_prompt(
                filename=filename, batch_text=content_text, category_list=cat_list,
            )
            messages = [
                {"role": "system", "content": get_prompt(session, "card_system", CARD_SYSTEM_PROMPT)},
                {"role": "user", "content": user_prompt},
            ]

            result = _run_async(ai.chat_completion(messages, temperature=0.2, feature="smart_import"))
            update_job_status(job_id, "running", progress=70)

            content = result["content"]
            if "```json" in content:
                content = content.split("```json")[1].split("```")[0]
            elif "```" in content:
                content = content.split("```")[1].split("```")[0]
            import re
            content = re.sub(r",\s*([}\]])", r"\1", content)
            cards_data = _json.loads(content)

            if isinstance(cards_data, dict) and "cards" in cards_data:
                cards_data = cards_data["cards"]

            created = 0
            for item in cards_data:
                distractors = item.get("distractors", [])
                if isinstance(distractors, list):
                    distractors_str = _json.dumps(distractors, ensure_ascii=False)
                else:
                    distractors_str = str(distractors) if distractors else ""
                meta_info = item.get("meta_info", "")
                if isinstance(meta_info, dict):
                    meta_info = _json.dumps(meta_info, ensure_ascii=False)
                cat_name = item.get("category", "")
                category = cat_map.get(cat_name)
                card = Card(
                    deck_id=deck_id, category_id=category.id if category else None,
                    front=item.get("front", ""), back=item.get("back", ""),
                    explanation=item.get("explanation", ""), distractors=distractors_str,
                    tags=item.get("tags", ""), meta_info=meta_info, is_ai_generated=True,
                )
                session.add(card)
                created += 1

            deck.card_count += created
            session.add(deck)
            session.commit()

            update_job_status(
                job_id, "completed",
                result_json=_json.dumps({"created": created, "total_parsed": len(cards_data)}, ensure_ascii=False),
            )
    except Exception as e:
        logger.error("BG smart import error: %s", e, exc_info=True)
        update_job_status(job_id, "failed", error_message=str(e)[:2000])


def _bg_batch_enrich(job_id: int, user_id: int, card_ids: list[int], deck_id: int | None, batch_size: int):
    """Background task for batch enrich."""
    from sqlmodel import Session as SyncSession
    from app.models.card import Card
    import json as _json

    update_job_status(job_id, "running", progress=5)

    try:
        with SyncSession(engine) as session:
            ai = AIService(session, user_id)
            if not ai.is_available():
                update_job_status(job_id, "failed", error_message=ai.get_unavailable_reason())
                return

            if card_ids:
                cards_to_enrich = [session.get(Card, cid) for cid in card_ids]
                cards_to_enrich = [c for c in cards_to_enrich if c]
            elif deck_id:
                cards_to_enrich = list(session.exec(select(Card).where(Card.deck_id == deck_id)).all())
            else:
                update_job_status(job_id, "failed", error_message="需要提供 card_ids 或 deck_id")
                return

            total = len(cards_to_enrich)
            enriched = 0
            skipped = 0
            errors = 0

            for i in range(0, total, batch_size):
                batch = cards_to_enrich[i:i + batch_size]
                progress = int(10 + 80 * i / max(total, 1))
                update_job_status(job_id, "running", progress=progress)

                cards_text = []
                for idx, card in enumerate(batch):
                    has_distractors = bool(card.distractors and card.distractors.strip() and card.distractors != "[]")
                    has_meta = bool(card.meta_info and card.meta_info.strip() and card.meta_info not in ("{}", ""))
                    if has_distractors and has_meta:
                        skipped += 1
                        continue
                    cards_text.append({
                        "index": idx, "card_id": card.id,
                        "front": card.front[:200], "back": card.back[:200],
                        "explanation": (card.explanation or "")[:100],
                        "needs_distractors": not has_distractors, "needs_meta": not has_meta,
                    })

                if not cards_text:
                    continue

                prompt_cards = _json.dumps(cards_text, ensure_ascii=False)
                default_enrich_prompt = get_prompt(session, "batch_enrich",
                    "你是出题专家。请为以下卡片补充缺失的信息。\n"
                    "对于needs_distractors=true的卡片，生成3个合理的干扰项(distractors)。\n"
                    "对于needs_meta=true的卡片，生成meta_info。\n"
                    "回复JSON数组: [{\"card_id\": 1, \"distractors\": [\"A\",\"B\",\"C\"], \"meta_info\": {...}}, ...]"
                )
                messages = [
                    {"role": "system", "content": default_enrich_prompt},
                    {"role": "user", "content": prompt_cards},
                ]

                try:
                    result = _run_async(ai.chat_completion(messages, temperature=0.3, feature="batch_enrich"))
                    content = result["content"]
                    if "```json" in content:
                        content = content.split("```json")[1].split("```")[0]
                    elif "```" in content:
                        content = content.split("```")[1].split("```")[0]
                    enrichments = _json.loads(content)

                    card_map = {c.id: c for c in batch}
                    for item in enrichments:
                        cid = item.get("card_id")
                        card = card_map.get(cid)
                        if not card:
                            continue
                        if "distractors" in item and item["distractors"]:
                            card.distractors = _json.dumps(item["distractors"], ensure_ascii=False)
                        if "meta_info" in item and item["meta_info"]:
                            meta = item["meta_info"]
                            if isinstance(meta, dict):
                                meta = _json.dumps(meta, ensure_ascii=False)
                            card.meta_info = meta
                        session.add(card)
                        enriched += 1
                    session.commit()
                except Exception as e:
                    logger.error("BG batch enrich batch error: %s", e)
                    errors += len(cards_text)

            update_job_status(
                job_id, "completed",
                result_json=_json.dumps({"total": total, "enriched": enriched, "skipped": skipped, "errors": errors}, ensure_ascii=False),
            )
    except Exception as e:
        logger.error("BG batch enrich error: %s", e, exc_info=True)
        update_job_status(job_id, "failed", error_message=str(e)[:2000])


def _bg_complete_and_create_cards(
    job_id: int, user_id: int, deck_id: int, category_id: int | None,
    cards_input: list[dict],
):
    """Background task: AI-complete cards then bulk create them."""
    from sqlmodel import Session as SyncSession
    from app.models.card import Card
    from app.models.deck import Deck
    from app.models.category import Category
    from app.services.prompts import CARD_SYSTEM_PROMPT
    import json as _json

    update_job_status(job_id, "running", progress=10)

    try:
        with SyncSession(engine) as session:
            ai = AIService(session, user_id)
            if not ai.is_available():
                update_job_status(job_id, "failed", error_message=ai.get_unavailable_reason())
                return

            deck = session.get(Deck, deck_id)
            if not deck:
                update_job_status(job_id, "failed", error_message="牌组不存在")
                return

            # Separate cards that need AI completion from those that don't
            needs_ai = [c for c in cards_input if not c.get("back", "").strip()]
            has_back = [c for c in cards_input if c.get("back", "").strip()]

            ai_results: dict[int, dict] = {}  # index -> completion

            if needs_ai:
                update_job_status(job_id, "running", progress=20)

                all_cats = session.exec(select(Category).where(Category.is_active == True)).all()
                cat_list = "、".join(c.name for c in all_cats)

                cards_text = [{"index": i, "front": c.get("front", ""), "category": c.get("category", "")} for i, c in enumerate(needs_ai)]
                user_prompt = (
                    f"请为以下卡片生成完整的内容（JSON数组）。\n"
                    f"每张卡片已有front（题目）和category（分类），请根据system prompt中的格式要求补充所有字段：\n"
                    f"back、explanation、distractors（3个）、tags、meta_info（含knowledge、exam_focus、alternate_questions、facts）。\n\n"
                    f"可用分类：{cat_list}\n\n"
                    f"待补全的卡片：\n{_json.dumps(cards_text, ensure_ascii=False)}\n\n"
                    f"回复纯JSON数组，每项包含index字段和所有补全的字段。"
                )
                messages = [
                    {"role": "system", "content": get_prompt(session, "card_system", CARD_SYSTEM_PROMPT)},
                    {"role": "user", "content": user_prompt},
                ]
                try:
                    result = _run_async(ai.chat_completion(messages, feature="complete_cards"))
                    content = result["content"]
                    if "```json" in content:
                        content = content.split("```json")[1].split("```")[0]
                    elif "```" in content:
                        content = content.split("```")[1].split("```")[0]
                    import re
                    content = re.sub(r",\s*([}\]])", r"\1", content)
                    parsed = _json.loads(content)
                    if isinstance(parsed, dict):
                        completions = parsed.get("cards", [])
                    elif isinstance(parsed, list):
                        completions = parsed
                    else:
                        completions = []
                    for item in completions:
                        idx = item.get("index", -1)
                        if 0 <= idx < len(needs_ai):
                            ai_results[idx] = item
                except Exception as e:
                    logger.warning("BG complete cards AI error: %s", e)

            update_job_status(job_id, "running", progress=70)

            # Build and create cards
            created = 0
            for c in cards_input:
                front = c.get("front", "").strip()
                if not front:
                    continue

                # Check if this card was in needs_ai list
                ai_idx = None
                for i, na in enumerate(needs_ai):
                    if na is c:
                        ai_idx = i
                        break
                comp = ai_results.get(ai_idx, {}) if ai_idx is not None else {}

                back = c.get("back", "").strip() or comp.get("back", "")
                explanation = c.get("explanation", "").strip() or comp.get("explanation", "")
                distractors_raw = c.get("distractors", "").strip() or comp.get("distractors", [])
                if isinstance(distractors_raw, list):
                    distractors_str = _json.dumps(distractors_raw, ensure_ascii=False)
                elif isinstance(distractors_raw, str) and distractors_raw and not distractors_raw.startswith("["):
                    dlist = [s.strip() for s in distractors_raw.replace("，", ",").split(",") if s.strip()]
                    distractors_str = _json.dumps(dlist, ensure_ascii=False) if dlist else ""
                else:
                    distractors_str = str(distractors_raw) if distractors_raw else ""
                meta_info = comp.get("meta_info", "")
                if isinstance(meta_info, dict):
                    meta_info = _json.dumps(meta_info, ensure_ascii=False)
                tags = c.get("tags", "").strip() or comp.get("tags", "")

                card = Card(
                    deck_id=deck_id,
                    category_id=category_id or deck.category_id,
                    front=front,
                    back=back,
                    explanation=explanation,
                    distractors=distractors_str,
                    tags=tags,
                    meta_info=meta_info,
                    is_ai_generated=bool(comp),
                )
                session.add(card)
                created += 1

            deck.card_count = (deck.card_count or 0) + created
            session.add(deck)
            session.commit()

            update_job_status(
                job_id, "completed",
                result_json=_json.dumps({
                    "created": created,
                    "ai_enriched": len(ai_results),
                    "total": len(cards_input),
                }, ensure_ascii=False),
            )
    except Exception as e:
        logger.error("BG complete and create cards error: %s", e, exc_info=True)
        update_job_status(job_id, "failed", error_message=str(e)[:2000])


@router.post("/complete-cards/async")
async def complete_cards_async(
    data: dict,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """Async version: AI-complete cards + create them in background."""
    cards_input = data.get("cards", [])
    deck_id = data.get("deck_id")
    category_id = data.get("category_id")
    if not cards_input or not deck_id:
        raise HTTPException(status_code=400, detail="需要提供 cards 和 deck_id")

    job = create_job(session, current_user.id, "complete_cards", f"AI补全并创建 {len(cards_input)} 张卡片")
    background_tasks.add_task(
        _bg_complete_and_create_cards, job.id, current_user.id, deck_id, category_id, cards_input
    )
    return {"job_id": job.id, "message": f"AI 卡片创建任务已提交，{len(cards_input)} 张卡片后台处理中"}


@router.post("/smart-import/async")
async def smart_import_async(
    deck_id: int,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """Start smart import as a background job. Returns job ID immediately."""
    file_bytes = await file.read()
    filename = file.filename or "unknown"
    job = create_job(session, current_user.id, "smart_import", f"智能导入: {filename}")
    background_tasks.add_task(_bg_smart_import, job.id, current_user.id, deck_id, file_bytes, filename)
    return {"job_id": job.id, "message": "智能导入任务已创建，后台处理中"}


@router.post("/batch-enrich/async")
async def batch_enrich_async(
    data: dict,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_user),
    session: Session = Depends(get_session),
):
    """Start batch enrich as a background job."""
    card_ids = data.get("card_ids", [])
    deck_id = data.get("deck_id")
    batch_size = data.get("batch_size", 50)
    title = f"批量补充: {len(card_ids)} 张卡片" if card_ids else f"批量补充: 牌组 #{deck_id}"
    job = create_job(session, current_user.id, "batch_enrich", title)
    background_tasks.add_task(_bg_batch_enrich, job.id, current_user.id, card_ids, deck_id, batch_size)
    return {"job_id": job.id, "message": "批量补充任务已创建，后台处理中"}
